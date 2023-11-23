# 시각화 및 엑셀 저장을 위한 모듈
from tkinter import Y
import openpyxl
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import pandas as pd
import cv2
import math


class Preprocessing:
    def __init__(self):
        # 시각화 그래프 세팅
        """
        plt.ion()#실시간 업데이트 허용
        self.fig = plt.figure()
        self.ax = self.fig.add_subplot(111)#1개의 plot인 ax를 (1,1)지점에 구현
        plt.xlim(0,640)
        plt.ylim(0,480)
        plt.axis('square')
        """
        #3d
        #fig = plt.figure()
        #ax = fig.add_subplot(111, projection='3d')
        
        # camera input
        self.result_counter = 0  # 결과 파일명에 사용될 카운터 변수 초기화
        self.image_width = 640
        self.image_height = 480
        self.excel_path = "./landmarks.xlsx"

        # 추출한 Point들을 xlxs에 저장하기 위한 기본 세팅
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.cell(row=1, column=2).value = "x"
        self.sheet.cell(row=1, column=3).value = "y"
        self.sheet.cell(row=1, column=4).value = "z"
        self.x = []
        self.y = []
        self.z = []
        for i in range(0, 13, 1):
            self.x.append(0)
            self.y.append(0)
            self.z.append(0)
        
        self.top_pad = 50
        self.bottom_pad = 50
        self.left_pad = 50
        self.right_pad = 50

    def get_current_image(self):
        # 엑셀 파일에서 데이터 읽기
        df = pd.read_excel(self.excel_path)

        # 데이터를 5개 단위로 분할하여 평균값 계산
        num_points = len(df)
        chunk_size = 3
        start_idx = 0
        avg_points = []  # 각 구간의 평균값 저장
        
        def get_current_calculate_distance(point1, point2):
            return math.sqrt((point2[0] - point1[0])**2 + (point2[1] - point1[1])**2)
        
        threshold_distance = 100  # 선을 그리기 위한 최소 거리

        while start_idx < num_points:
            # 각 구간의 끝 인덱스 계산
            end_idx = min(start_idx + chunk_size, num_points)

            # 현재 구간의 데이터 가져오기
            chunk_df = df.iloc[start_idx:end_idx]

            # 현재 구간의 x, y 좌표 평균값 계산
            avg_x = chunk_df['x'].mean()
            avg_y = chunk_df['y'].mean()

            # 평균값을 리스트에 추가
            avg_points.append((avg_x, avg_y))
        
            # 다음 구간의 시작 인덱스 업데이트
            start_idx += chunk_size
    
        # 그래프 생성
        plt.figure(figsize=(8, 6))
        plt.xlim(0, 640)
        plt.ylim(0, 480)
        plt.axis('square')
    
        # 데이터 포인트 그리기 (평균값)
        for i in range(len(avg_points) - 1):
            distance = get_current_calculate_distance(avg_points[i], avg_points[i + 1])
            if distance <= threshold_distance:  # 일정 거리 이하인 경우에만 선을 그립니다.
                plt.scatter(avg_points[i][0], avg_points[i][1], color='white', marker='o', s=100)  # 평균값으로 하나의 점 그리기
                plt.plot([avg_points[i][0], avg_points[i + 1][0]], [avg_points[i][1], avg_points[i + 1][1]], color='black', linewidth=5)  # 선으로 연결하기

        # 마지막 평균값은 점만 그리기
        plt.scatter(avg_points[-1][0], avg_points[-1][1], color='white', marker='o', s=100)
    
        # 데이터 포인트 그리기 (전체 데이터)
        #plt.scatter(df['x'], df['y'], color='red')  # 전체 데이터의 점 그리기
    
        # 축과 레이블 숨기기
        plt.axis('off')
    
        # x축 상하 반전
        plt.gca().invert_xaxis()

        # y축 좌우 반전
        plt.gca().invert_yaxis()
    
        # 이미지로 저장
        plt.savefig('number.png', bbox_inches='tight', pad_inches=0)
        plt.close()

    def get_current_streaming_camera(self, ax):
        # 5.2 실시간 주요포인트 시연
        ax.clear()# 이전에 그린 plot 초기화
        plt.xlim(0,640)
        plt.ylim(0,480)
        plt.scatter(x, y)# 산점도 plot에 그리기
        plt.draw()# 현재 plot 시각화
        """
        #3d 그래프 실시간 업데이트 구현
        ax.clear()
        ax.scatter(x, y, z, c='r', marker='o')
        ax.set_xlabel('X Label')
        ax.set_ylabel('Y Label')
        ax.set_zlabel('Z Label')
        ax.set_xlim([0, 0.7])  # x 축 고정 (범위 설정)
        ax.set_ylim([0, 0.7])  # y 축 고정 (범위 설정)
        ax.set_zlim([-0.2, 0.2])  # z 축 고정 (범위 설정)

        plt.draw()
        plt.pause(0.1)  # 업데이트 간격 설정 (여기서는 0.1초)
        """
    
    def get_current_excel(self, save_frames, a, b):
        for i in range(0, save_frames, 1):
            self.sheet.cell(row=i + 2, column=1).value = i
            self.sheet.cell(row=i + 2, column=2).value = 640*a[i]
            self.sheet.cell(row=i + 2, column=3).value = 480*b[i]
        self.workbook.save('landmarks.xlsx')   
    
    def get_current_roi(self, binary_image):
        x_min = float('inf')
        x_max = float('-inf')
        y_min = float('inf')
        y_max = float('-inf')

        height, width = binary_image.shape[:2]

        for i in range(width):
            for j in range(height):
                if binary_image[j][i] == 255:  # 여기서 특정값은 검정색이라고 가정
                    if i < x_min:
                        x_min = i
                    if i > x_max:
                        x_max = i
                    if j < y_min:
                        y_min = j
                    if j > y_max:
                        y_max = j

        # ROI 추출
        roi = binary_image[y_min:y_max, x_min:x_max]

        return x_min, y_min, x_max-x_min, y_max-y_min, roi

    def get_current_resize(self):
        # 바이너리 이미지 불러오기 (흰색 영역에 대한 윤곽선을 찾습니다)
        color_img = cv2.imread('number.png')
        binary_img = cv2.imread('number.png', cv2.IMREAD_GRAYSCALE)  # 바이너리 이미지 경로를 지정해주세요.

        # 바이너리 이미지를 반전시킵니다. (검은색과 흰색 반전)
        binary_img_inverted = cv2.bitwise_not(binary_img)

        # 검은색 영역에서 가장 큰 영역의 ROI와 위치 정보 찾기
        x, y, w, h, black_roi = self.get_current_roi(binary_img_inverted)

        original_roi = color_img[y:y+h, x:x+w]
        padded_image_white_bg = self.get_current_padding(original_roi, self.top_pad, self.bottom_pad, self.left_pad, self.right_pad)
        # 원하는 크기로 resize
        desired_width = 224
        desired_height = 224
        resized_roi = cv2.resize(padded_image_white_bg, (desired_width, desired_height))

        # 이미지를 저장할 경로와 파일명 설정
        save_path = f'./Result/Result_{self.result_counter}.jpg'
        # 이미지 저장
        cv2.imwrite(save_path, resized_roi)

        self.result_counter += 1

    def get_current_padding(self, image, top, bottom, left, right):
        padded_image = cv2.copyMakeBorder(image, top, bottom, left, right, cv2.BORDER_CONSTANT, value=(255, 255, 255))
        return padded_image

        
