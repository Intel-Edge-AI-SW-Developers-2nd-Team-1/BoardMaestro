################################################################################################
#사용할 모듈
#사용전에 꼭 cv2, numpy, mediapipe 모듈을 설치하여야 구동이 됩니다.
from tkinter.tix import Tree
import cv2
import numpy as np

import time

#시각화 및 엑셀 저장을 위한 모듈
import openpyxl
import matplotlib.pyplot as plt
import pandas as pd

#미디어파이프를 사용하기 위한 모듈
#python -m pip install mediapipe
import mediapipe as mp
from mediapipe.tasks import python
from mediapipe.tasks.python import vision

from mediapipe import solutions
from mediapipe.framework.formats import landmark_pb2

#정의할 상수
MARGIN = 10  # pixels
FONT_SIZE = 1
FONT_THICKNESS = 1
HANDEDNESS_TEXT_COLOR = (88, 205, 54) # vibrant green

###########################################################################################################
model_path = '/home/ubuntu/intel-project/models/hand_landmarker.task'

BaseOptions = mp.tasks.BaseOptions
HandLandmarker = mp.tasks.vision.HandLandmarker
HandLandmarkerOptions = mp.tasks.vision.HandLandmarkerOptions
HandLandmarkerResult = mp.tasks.vision.HandLandmarkerResult
VisionRunningMode = mp.tasks.vision.RunningMode

############################################################################################################
# 사용할 함수 정의

#결과 출력
def print_result(result: HandLandmarkerResult, output_image: mp.Image, timestamp_ms: int):
    '''Create a hand landmarker instance with the live stream mode:'''
    print('hand landmarker result: {}'.format(result))

#이미지에 랜드마크 추출하는 함수
def draw_landmarks_on_image(rgb_image, detection_result):
    '''to visualize the hand landmark detection results'''
    hand_landmarks_list = detection_result.hand_landmarks
    handedness_list = detection_result.handedness
    annotated_image = np.copy(rgb_image)

    # Loop through the detected hands to visualize.
    for idx in range(len(hand_landmarks_list)):
        hand_landmarks = hand_landmarks_list[idx]
        handedness = handedness_list[idx]

        # Draw the hand landmarks.
        hand_landmarks_proto = landmark_pb2.NormalizedLandmarkList()
        hand_landmarks_proto.landmark.extend([
            landmark_pb2.NormalizedLandmark(x=landmark.x, y=landmark.y, z=landmark.z) for landmark in hand_landmarks])
        solutions.drawing_utils.draw_landmarks(
            annotated_image,
            hand_landmarks_proto,
            solutions.hands.HAND_CONNECTIONS,
            solutions.drawing_styles.get_default_hand_landmarks_style(),
            solutions.drawing_styles.get_default_hand_connections_style())

        # Get the top left corner of the detected hand's bounding box.
        height, width, _ = annotated_image.shape
        x_coordinates = [landmark.x for landmark in hand_landmarks]
        y_coordinates = [landmark.y for landmark in hand_landmarks]
        text_x = int(min(x_coordinates) * width)
        text_y = int(min(y_coordinates) * height) - MARGIN

        # Draw handedness (left or right hand) on the image.
        cv2.putText(annotated_image, f"{handedness[0].category_name}",
                    (text_x, text_y), cv2.FONT_HERSHEY_DUPLEX,
                    FONT_SIZE, HANDEDNESS_TEXT_COLOR, FONT_THICKNESS, cv2.LINE_AA)
    return annotated_image

#이미지를 plot을 이용해 그려서 숫자를 만들어주는 함수
def save_lines_without_axes(excel_path):
    # 엑셀 파일에서 데이터 읽기
    df = pd.read_excel(excel_path)
    
    # 데이터를 5개 단위로 분할하여 평균값 계산
    num_points = len(df)
    chunk_size = 4
    start_idx = 0
    avg_points = []  # 각 구간의 평균값 저장
    
    while start_idx < num_points:
        # 각 구간의 끝 인덱스 계산
        end_idx = min(start_idx + chunk_size, num_points)
        
        # 현재 구간의 데이터 가져오기
        chunk_df = df.iloc[start_idx:end_idx]
        
        # 현재 구간의 x, y 좌표 평균값 계산
        avg_x = chunk_df['x'].mean()
        avg_y = chunk_df['y'].mean()
        
        # 평균값을 리스트에 추가
        avg_points.append((avg_x, avg_y))
        
        # 다음 구간의 시작 인덱스 업데이트
        start_idx += chunk_size
    
    # 그래프 생성
    plt.figure(figsize=(8, 6))
    
    # 데이터 포인트 그리기 (평균값)
    for i in range(len(avg_points) - 1):
        plt.scatter(avg_points[i][0], avg_points[i][1], color='white', marker='o', s=100)  # 평균값으로 하나의 점 그리기
        plt.plot([avg_points[i][0], avg_points[i + 1][0]], [avg_points[i][1], avg_points[i + 1][1]], color='black', linewidth=10)  # 선으로 연결하기

    
    # 마지막 평균값은 점만 그리기
    plt.scatter(avg_points[-1][0], avg_points[-1][1], color='white', marker='o', s=100)
    
    # 데이터 포인트 그리기 (전체 데이터)
    plt.scatter(df['x'], df['y'], color='red')  # 전체 데이터의 점 그리기
    
    # 축과 레이블 숨기기
    plt.axis('off')
    
    # x축 상하 반전
    plt.gca().invert_xaxis()

    # y축 좌우 반전
    plt.gca().invert_yaxis()
    
    # 이미지로 저장
    plt.savefig('lines_without_axes_plot_all.png', bbox_inches='tight', pad_inches=0)
    plt.close()

#xlxs 값에 저장을 위한 함수
#def write_to_xlxs():
####################################################################################################  
# set option
options = HandLandmarkerOptions(
    base_options=BaseOptions(model_asset_path=model_path, delegate=BaseOptions.Delegate.GPU),
    running_mode=VisionRunningMode.IMAGE)

# The landmarker is initialized
landmarker = HandLandmarker.create_from_options(options)

# Use OpenCV’s VideoCapture to start capturing from the webcam.
cap = cv2.VideoCapture(0)

#추출한 Point들을 xlxs에 저장하기 위한 기본 세팅
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=2).value = "x"
sheet.cell(row=1, column=3).value = "y"
sheet.cell(row=1, column=4).value = "z"
x=[]
y=[]
z=[]
for i in range(0,13,1):
    x.append(0)
    y.append(0)
    z.append(0)

#시각화 그래프 세팅
plt.ion()#실시간 업데이트 허용
fig = plt.figure()
ax = fig.add_subplot(111)#1개의 plot인 ax를 (1,1)지점에 구현

#사용할 지점
position = [
    "0 - wrist",
    "1 - thumb_cmc",
    "2 - thumb_mcp",
    "3 - thumb_ip",
    "4 - thumb_tip",
    "5 - index_finger_mcp",
    "6 - index_finger_pip",
    "7 - index_finger_dip",
    "8 - index_finger_tip",
    "9 - middle_finger_mcp",
    "10 - middle_finger_pip",
    "11 - middle_finger_dip",
    "12 - middle_finger_tip",
    "13 - ring_finger_mcp",
    "14 - ring_finger_pip",
    "15 - ring_finger_dip",
    "16 - ring_finger_tip",
    "17 - pinky_finger_mcp",
    "18 - pinky_finger_pip",
    "19 - pinky_finger_dip",
    "20 - pinky_finger_tip",   
]

landmark_indices = [8]

#저장될 값을 위한 flag 셋팅
status = False
flag = False
start_frame = 0

#STEP Pre normalization before (정규화 된 좌표를 이미지 좌표로 전환) - 카메라 기본 셋팅
image_width = 640
image_height = 480

######################################################################################################
#while 문 시작
# Create a loop to read the latest frame from the camera using VideoCapture#read()
while True:
    # to calculate running time, save start time to start_time
    start_time = time.time()

    # read the frame from webcam
    ret, frame = cap.read()

    # check the frame
    if not ret:
        break

    # Convert the frame received from OpenCV to a MediaPipe’s Image object.
    mp_image = mp.Image(image_format = mp.ImageFormat.SRGB,
                        data = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))

    '''
    # Send live image data to perform hand landmarks detection.
    # The results are accessible via the `result_callback` provided in
    # the `HandLandmarkerOptions` object.
    # The hand landmarker must be created with the live stream mode.
    landmarker.detect_async(mp_image, time.time_ns() // 1_000_000)
    '''
    hand_landmarker_result = landmarker.detect(mp_image)
    
    # show detection_result to visible
    annotated_image = draw_landmarks_on_image(mp_image.numpy_view(), hand_landmarker_result)
    cv2.imshow("Result", cv2.cvtColor(annotated_image, cv2.COLOR_RGB2BGR))


    if(hand_landmarker_result.hand_landmarks.__len__() != 1):continue #사람이 탐색 되지않으면 아랫부분 시행 중지
    
    # 지금 오른쪽 손목 기준 좌표 그림 추출
    if(status == True):
        for i in range(len(landmark_indices)):
            index = landmark_indices[i]
            sheet.cell(row=start_frame + 2, column=1).value = start_frame
            sheet.cell(row=start_frame + 2, column=2).value = (image_width)*(hand_landmarker_result.hand_landmarks[0][index].x)
            sheet.cell(row=start_frame + 2, column=3).value = (image_height)*(hand_landmarker_result.hand_landmarks[0][index].y)
            sheet.cell(row=start_frame + 2, column=4).value = (image_height*image_width)*(hand_landmarker_result.hand_landmarks[0][index].z)
            x[i] = image_width * (hand_landmarker_result.hand_landmarks[0][index].x)
            y[i] = image_height * (hand_landmarker_result.hand_landmarks[0][index].y)
            z[i] = (hand_landmarker_result.hand_landmarks[0][index].z)
            #sheet.cell(row=start_frame + 2, column=4).value = detection_result.pose_landmarks[0][16].z
        start_frame += 1
        workbook.save('landmarks.xlsx')    
    
    if(flag == True):
        save_lines_without_axes('./landmarks.xlsx')
        flag = False

    #5.2 실시간 주요포인트 시연
    ax.clear()#이전에 그린 plot 초기화
    plt.xlim(0,640)
    plt.ylim(0,480)
    plt.scatter(x,y)#산점도 plot에 그리기
    plt.draw()#현재 plot 시각화
    plt.pause(0.005)#시각화한 plot을 보여주기위해 잠깐대기
    plt.show()


    # show process time and fps
    process_time = time.time() - start_time
    FPS = 1 / process_time
    print(f"process_time = {process_time:.4f}s, FPS = {FPS:.2f}")


    # if pressed 'q', end capture
    key = cv2.waitKey(1) & 0xFF
    if key == ord('q'):
        break
    elif key == ord('s'):
        status = not status
        print(status)
    elif key == ord('c'):
        flag = not flag
        print(flag)    

# 카메라 객체와 윈도우 창 닫기
cap.release()
cv2.destroyAllWindows()

###########################################################################################